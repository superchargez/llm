import os; os.chdir(r"../excel_files")
import polars as pl
from openpyxl import load_workbook

def load_sheets_as_polars(file_path):
    wb = load_workbook(file_path, data_only=True)
    sheet_dataframes = {}
    
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        
        # Extract rows as lists of values
        data = [[cell.value for cell in row] for row in sheet.iter_rows()]
        
        # Create a Polars DataFrame with `strict=False` to handle mixed types
        df = pl.DataFrame(data, strict=False)
        df = df.drop_nulls()  # Drops fully null rows/columns
        sheet_dataframes[sheet_name] = df
    
    return sheet_dataframes

def detect_tables(df, max_empty_rows=2):
    """
    Detect possible tables in a Polars DataFrame based on clusters of populated rows.
    
    :param df: Polars DataFrame representing a sheet.
    :param max_empty_rows: Maximum number of consecutive empty rows allowed within a table.
    :return: List of detected tables (as Polars DataFrames).
    """
    tables = []
    current_table = []
    empty_row_count = 0

    for row in df.rows():
        non_empty_cells = sum(cell is not None for cell in row)
        
        if non_empty_cells > 0:  # Row has data
            # If there were any previous empty rows, reset the count and continue the current table
            empty_row_count = 0
            current_table.append(row)
        else:  # Row is empty
            empty_row_count += 1
            # If the empty row count exceeds `max_empty_rows`, finalize the current table
            if empty_row_count > max_empty_rows and current_table:
                tables.append(pl.DataFrame(current_table, strict=False))
                current_table = []  # Start a new table after gap

    # Append any remaining data as the last table
    if current_table:
        tables.append(pl.DataFrame(current_table, strict=False))
    
    return tables

def extract_tables(file_path):
    sheets = load_sheets_as_polars(file_path)
    all_tables = []

    for sheet_name, df in sheets.items():
        tables = detect_tables(df)
        for table in tables:
            all_tables.append((sheet_name, table))

    return all_tables

# Use this function to process a given Excel file
file_path = 'ticket_sales.xlsx'
tables = extract_tables(file_path)

# Output each table as needed
for sheet_name, table in tables:
    print(f"Sheet: {sheet_name}")
    print(table)
