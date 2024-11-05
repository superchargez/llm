from langchain.tools import ExcelTool
from openpyxl import load_workbook
from ToolAgents.agents import MistralAgent
from ToolAgents.provider import LlamaCppServerProvider
import socket

class ExcelInspectorTool:
    def _call(self, query: str) -> dict:
        file_path = query.split(" ")[1]  # Assuming the query is "Analyze <file_path>"
        return excel_inspector(file_path)

def excel_inspector(file_path):
    wb = load_workbook(filename=file_path)
    sheets = wb.sheetnames
    
    sheet_info = []
    for sheet_name in sheets:
        ws = wb[sheet_name]
        
        # Simple table detection (you might need to enhance this)
        tables = []
        for i in range(1, ws.max_row + 1):
            if ws.cell(row=i, column=1).value:
                table = []
                for j in range(1, ws.max_column + 1):
                    if ws.cell(row=i, column=j).value:
                        table.append(ws.cell(row=i, column=j).value)
                    else:
                        break
                tables.append(table)
        
        columns = [] if len(tables) == 0 else [str(cell) for cell in tables[0]]
        dtypes = ['str'] * len(columns)  # Simplified type detection
        
        sheet_info.append({
            "sheet_name": sheet_name,
            "tables": len(tables),
            "columns": columns,
            "dtypes": dtypes
        })
    
    return {
        "sheets": len(sheets),
        "sheet_names": sheets,
        "sheet_info": sheet_info
    }

# Create a custom tool for our Excel inspector
class ExcelInspectorTool(ExcelTool):
    def _call(self, query: str) -> dict:
        file_path = query.split(" ")[1]  # Assuming the query is "Analyze <file_path>"
        return excel_inspector(file_path)

# Initialize the agent with our custom tool
provider = LlamaCppServerProvider(f"http://{socket.gethostname()}.local:8384")
agent = MistralAgent(provider=provider, debug_output=False)

def run_excel_agent(prompt):
    return agent.run(query=prompt)

PROMPT = """You are an Excel inspector. Your task is to analyze the given Excel file and report on the following:
1. Number of sheets in the workbook
2. List of sheet names
3. For each sheet:
   - Number of tables found
   - Column names and data types for each table

Please respond in a structured format."""

file_path = "../../excel_files/tickes_sales.xlsx"
query = f"Analyze {file_path}"

result = run_excel_agent(query)
print(result)
