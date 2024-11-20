from typing import Optional, List, Dict, Any
import json
import re
from pathlib import Path
import pandas as pd
import openpyxl
from fastapi import FastAPI
from pydantic import BaseModel
from jinja2 import Template
import torch
from transformers import AutoModelForCausalLM, AutoTokenizer, BitsAndBytesConfig
import spacy
import logging
from dataclasses import dataclass, field
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

app = FastAPI()

system_prompt = Template("""You are an expert in composing functions. You are given a question and a set of possible functions. 
Based on the question, you will need to make one or more function/tool calls to achieve the purpose. 
If none of the functions can be used, point it out and refuse to answer. 
If the given question lacks the parameters required by the function, also point it out.

You have access to the following tools:
<tools>{{ tools }}</tools>

The output MUST strictly adhere to the following format, and NO other text MUST be included.
The example format is as follows. Please make sure the parameter type is correct. If no function call is needed, please make the tool calls an empty list '[]'.
<tool_call>[
{"name": "func_name1", "arguments": {"argument1": "value1", "argument2": "value2"}},
... (more tool calls as required)
]</tool_call>""")

def prepare_messages(
    query: str,
    tools: Optional[dict[str, Any]] = None,
    history: Optional[List[Dict[str, str]]] = None
) -> List[Dict[str, str]]:
    """Prepare the system and user messages for the given query and tools.
    
    Args:
        query: The query to be answered.
        tools: The tools available to the user. Defaults to None, in which case if a
            list without content will be passed to the model.
        history: Exchange of messages, including the system_prompt from
            the first query. Defaults to None, the first message in a conversation.
    """
    if tools is None:
        tools = []
    if history:
        messages = history.copy()
        messages.append({"role": "user", "content": query})
    else:
        messages = [
            {"role": "system", "content": system_prompt.render(tools=json.dumps(tools))},
            {"role": "user", "content": query}
        ]
    return messages

def parse_response(text: str) -> str | Dict[str, Any]:
    """Parses a response from the model, returning either the
    parsed list with the tool calls parsed, or the
    model thought or response if couldn't generate one.

    Args:
        text: Response from the model.
    """
    pattern = r"<tool_call>(.*?)</tool_call>"
    matches = re.findall(pattern, text, re.DOTALL)
    if matches:
        return json.loads(matches[0])
    return text

model_name_smollm = "HuggingFaceTB/SmolLM2-1.7B-Instruct"

# Use BitsAndBytesConfig for quantization
nf4_config = BitsAndBytesConfig(
    load_in_4bit=True,
    bnb_4bit_quant_type="nf4",
    bnb_4bit_use_double_quant=True,
    bnb_4bit_compute_dtype=torch.bfloat16
)

model = AutoModelForCausalLM.from_pretrained(
    model_name_smollm,
    quantization_config=nf4_config,
    device_map="auto",
    torch_dtype=torch.float16,  # Use float16 for mixed precision training
    low_cpu_mem_usage=True,  # Enable CPU memory optimization
    trust_remote_code=True
)

tokenizer = AutoTokenizer.from_pretrained(model_name_smollm)

@dataclass
class WorksheetContent:
    name: str
    content_type: str = None
    data_regions: List[Dict[str, Any]] = field(default_factory=list)
    charts: List[Dict[str, Any]] = field(default_factory=list)
    unstructured_text: List[Dict[str, Any]] = field(default_factory=list)
    metadata: Dict[str, Any] = field(default_factory=dict)

class ContentType:
    STRUCTURED_TABLE = "structured_table"
    UNSTRUCTURED_TEXT = "unstructured_text"
    CHART = "chart"
    MIXED = "mixed"
    EMPTY = "empty"

class ExcelAnalyzer:
    """Analyzes Excel content and determines its structure and type."""
    
    def detect_content_type(self, ws: openpyxl.worksheet.worksheet.Worksheet) -> str:
        """Detect the type of content in a worksheet."""
        has_tables = False
        has_text = False
        has_charts = False
        
        # Check for charts
        if ws._charts:
            has_charts = True
            
        # Analyze cells for content type
        data_regions = self._find_data_regions(ws)
        text_regions = self._find_text_regions(ws)
        
        if data_regions:
            has_tables = True
        if text_regions:
            has_text = True
            
        if has_tables and has_text:
            return ContentType.MIXED
        elif has_tables:
            return ContentType.STRUCTURED_TABLE
        elif has_text:
            return ContentType.UNSTRUCTURED_TEXT
        elif has_charts:
            return ContentType.CHART
        else:
            return ContentType.EMPTY

    def _find_data_regions(self, ws) -> List[Dict[str, Any]]:
        """Find regions containing tabular data."""
        regions = []
        current_region = None
        
        for row in ws.iter_rows():
            row_values = [cell.value for cell in row]
            if any(row_values):  # Non-empty row
                if self._is_header_row(row_values):
                    if current_region:
                        regions.append(current_region)
                    current_region = {
                        'start_row': row[0].row,
                        'headers': row_values,
                        'data': []
                    }
                elif current_region:
                    current_region['data'].append(row_values)
            elif current_region:
                regions.append(current_region)
                current_region = None
                
        if current_region:
            regions.append(current_region)
            
        return regions

    def _is_header_row(self, row_values: List[Any]) -> bool:
        """Determine if a row is likely a header row."""
        non_empty = [v for v in row_values if v is not None]
        if not non_empty:
            return False
        return all(isinstance(v, str) and len(v) < 100 for v in non_empty)

    def _find_text_regions(self, ws) -> List[Dict[str, Any]]:
        """Find regions containing unstructured text."""
        text_regions = []
        current_text = []
        
        for row in ws.iter_rows():
            cell_values = [cell.value for cell in row if cell.value]
            if cell_values:
                # Check if this looks like unstructured text
                if any(isinstance(v, str) and len(v) > 100 for v in cell_values):
                    current_text.extend(cell_values)
                elif current_text:
                    text_regions.append({
                        'content': '\n'.join(current_text),
                        'row': row[0].row
                    })
                    current_text = []
                    
        if current_text:
            text_regions.append({
                'content': '\n'.join(current_text),
                'row': ws.max_row
            })
            
        return text_regions

class SmartExcelManager:
    def __init__(self, excel_directory: str = "../excel_files"):
        self.excel_directory = Path(excel_directory)
        if not self.excel_directory.exists():
            self.excel_directory.mkdir(parents=True)
        self.file_cache = {}
        self.content_cache = {}
        self.analyzer = ExcelAnalyzer()
        logger.info(f"Initialized SmartExcelManager with directory: {excel_directory}")
        
    def list_files(self) -> List[str]:
        """List all Excel files in the directory."""
        return [f.name for f in self.excel_directory.glob("*.xlsx")]

    def analyze_file(self, filename: str) -> Dict[str, Any]:
        """Thoroughly analyze an Excel file and cache its content structure."""
        if filename in self.content_cache:
            return self.content_cache[filename]
        
        file_path = self.excel_directory / filename
        if not file_path.exists():
            raise FileNotFoundError(f"File not found: {filename}")
            
        logger.info(f"Analyzing file: {filename}")
        try:
            wb = openpyxl.load_workbook(file_path, data_only=True)
            analysis = {
                'worksheets': {},
                'metadata': {
                    'total_sheets': len(wb.sheetnames),
                    'has_charts': any(ws._charts for ws in wb.worksheets),
                    'file_size': file_path.stat().st_size
                }
            }
            
            for ws in wb.worksheets:
                content = WorksheetContent(ws.title)
                content.content_type = self.analyzer.detect_content_type(ws)
                
                if content.content_type in [ContentType.STRUCTURED_TABLE, ContentType.MIXED]:
                    content.data_regions = self.analyzer._find_data_regions(ws)
                    
                if content.content_type in [ContentType.UNSTRUCTURED_TEXT, ContentType.MIXED]:
                    content.unstructured_text = self.analyzer._find_text_regions(ws)
                    
                if ws._charts:
                    content.charts = [{'type': chart.type, 'anchor': chart.anchor} 
                                    for chart in ws._charts]
                    
                analysis['worksheets'][ws.title] = vars(content)
                
            self.content_cache[filename] = analysis
            return analysis
            
        except Exception as e:
            logger.error(f"Error analyzing file {filename}: {e}")
            raise

    def query_content(self, filename: str, query: str) -> Dict[str, Any]:
        """Query content from an Excel file."""
        try:
            file_path = self.excel_directory / filename
            if not file_path.exists():
                return {"error": f"File not found: {filename}"}

            # For simple column average query
            if 'average' in query.lower() and 'column' in query.lower():
                df = pd.read_excel(file_path)
                # Extract column name from query
                column_match = re.search(r"'([^']+)'", query)
                if column_match:
                    column_name = column_match.group(1)
                    if column_name in df.columns:
                        avg_value = df[column_name].mean()
                        return {
                            "type": "structured",
                            "data": {
                                "column": column_name,
                                "average": float(avg_value)
                            }
                        }
                    else:
                        return {"error": f"Column '{column_name}' not found"}

            # Analyze file content
            analysis = self.analyze_file(filename)
            return {
                "type": "analysis",
                "data": analysis
            }

        except Exception as e:
            logger.error(f"Error querying content: {e}")
            return {"error": str(e)}

class QueryRequest(BaseModel):
    query: str
    history: Optional[List[Dict[str, str]]] = None

class SmartExcelAgent:
    def __init__(self, excel_manager: SmartExcelManager):
        self.excel_manager = excel_manager
        logger.info("Initialized SmartExcelAgent")
        
    async def process_query(self, query: str) -> Dict[str, Any]:
        """Process a query about Excel files."""
        try:
            # List available files
            available_files = self.excel_manager.list_files()
            logger.info(f"Available files: {available_files}")
            
            if not available_files:
                return {
                    'status': 'error',
                    'message': 'No Excel files found in directory'
                }

            # Extract filename from query if present
            filename_match = re.search(r"of\s+([^\s?]+\.xlsx)", query)
            if filename_match:
                filename = filename_match.group(1)
                if filename in available_files:
                    result = self.excel_manager.query_content(filename, query)
                    return {
                        'status': 'success',
                        'results': [{'file': filename, 'result': result}]
                    }
                else:
                    return {
                        'status': 'error',
                        'message': f'File {filename} not found'
                    }

            # If no specific file mentioned, search all files
            results = []
            for file in available_files:
                result = self.excel_manager.query_content(file, query)
                if 'error' not in result:
                    results.append({
                        'file': file,
                        'result': result
                    })

            return {
                'status': 'success',
                'results': results,
                'files_checked': len(available_files)
            }

        except Exception as e:
            logger.error(f"Error processing query: {e}")
            return {
                'status': 'error',
                'message': str(e)
            }

# Initialize the system
excel_manager = SmartExcelManager()
agent = SmartExcelAgent(excel_manager)

@app.post("/query")
async def query_model(request: QueryRequest):
    logger.info(f"Received query: {request.query}")
    return await agent.process_query(request.query)