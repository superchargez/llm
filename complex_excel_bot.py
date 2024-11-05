from typing import Optional, List, Dict, Any, Union, Tuple
import json
import re
from pathlib import Path
import pandas as pd
import numpy as np
import openpyxl
from openpyxl.chart import Chart
from openpyxl.drawing.image import Image
# import pytesseract  # For extracting text from images if needed
from PIL import Image as PILImage
import io
from fastapi import FastAPI, HTTPException
from pydantic import BaseModel, Field
from jinja2 import Template
import torch
from transformers import AutoModelForCausalLM, AutoTokenizer, BitsAndBytesConfig
from datetime import datetime
import random
from typing_extensions import Literal
import spacy  # For text analysis
# from nltk.tokenize import sent_tokenize
import logging

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

from transformers.utils import get_json_schema

app = FastAPI()

def prepare_messages(
    query: str,
    tools: Optional[dict[str, Any]] = None,
    history: Optional[List[Dict[str, str]]] = None
) -> List[Dict[str, str]]:
    """Prepare the system and user messages for the given query and tools.
    
    Args:
        query: The query to be answered.
        tools: The tools available to the user. Defaults to None, in which case if a
            list without content will be passed to the model.
        history: Exchange of messages, including the system_prompt from
            the first query. Defaults to None, the first message in a conversation.
    """
    if tools is None:
        tools = []
    if history:
        messages = history.copy()
        messages.append({"role": "user", "content": query})
    else:
        messages = [
            {"role": "system", "content": system_prompt.render(tools=json.dumps(tools))},
            {"role": "user", "content": query}
        ]
    return messages

def parse_response(text: str) -> str | Dict[str, Any]:
    """Parses a response from the model, returning either the
    parsed list with the tool calls parsed, or the
    model thought or response if couldn't generate one.

    Args:
        text: Response from the model.
    """
    pattern = r"<tool_call>(.*?)</tool_call>"
    matches = re.findall(pattern, text, re.DOTALL)
    if matches:
        return json.loads(matches[0])
    return text

model_name_smollm = "HuggingFaceTB/SmolLM2-1.7B-Instruct"

# Use BitsAndBytesConfig for quantization
nf4_config = BitsAndBytesConfig(
    load_in_4bit=True,
    bnb_4bit_quant_type="nf4",
    bnb_4bit_use_double_quant=True,
    bnb_4bit_compute_dtype=torch.bfloat16
)

model = AutoModelForCausalLM.from_pretrained(
    model_name_smollm,
    quantization_config=nf4_config,
    device_map="auto",
    torch_dtype=torch.float16,  # Use float16 for mixed precision training
    low_cpu_mem_usage=True,  # Enable CPU memory optimization
    trust_remote_code=True
)

tokenizer = AutoTokenizer.from_pretrained(model_name_smollm)

class ContentType:
    STRUCTURED_TABLE = "structured_table"
    UNSTRUCTURED_TEXT = "unstructured_text"
    CHART = "chart"
    MIXED = "mixed"
    EMPTY = "empty"

class WorksheetContent:
    def __init__(self, name: str):
        self.name = name
        self.content_type: str = None
        self.data_regions: List[Dict[str, Any]] = []
        self.charts: List[Dict[str, Any]] = []
        self.unstructured_text: List[Dict[str, Any]] = []
        self.metadata: Dict[str, Any] = {}

class ExcelAnalyzer:
    """Analyzes Excel content and determines its structure and type."""
    
    def __init__(self):
        # Load NLP model for text analysis
        try:
            self.nlp = spacy.load("en_core_web_sm")
        except:
            self.nlp = None
            logger.warning("Spacy model not available. Text analysis will be limited.")

    def detect_content_type(self, ws: openpyxl.worksheet.worksheet.Worksheet) -> str:
        """Detect the type of content in a worksheet."""
        has_tables = False
        has_text = False
        has_charts = False
        
        # Check for charts
        if ws._charts:
            has_charts = True
            
        # Analyze cells for content type
        data_regions = self._find_data_regions(ws)
        text_regions = self._find_text_regions(ws)
        
        if data_regions:
            has_tables = True
        if text_regions:
            has_text = True
            
        if has_tables and has_text:
            return ContentType.MIXED
        elif has_tables:
            return ContentType.STRUCTURED_TABLE
        elif has_text:
            return ContentType.UNSTRUCTURED_TEXT
        elif has_charts:
            return ContentType.CHART
        else:
            return ContentType.EMPTY

    def _find_data_regions(self, ws) -> List[Dict[str, Any]]:
        """Find regions containing tabular data."""
        regions = []
        current_region = None
        
        for row in ws.iter_rows():
            row_values = [cell.value for cell in row]
            if any(row_values):  # Non-empty row
                if self._is_header_row(row_values):
                    if current_region:
                        regions.append(current_region)
                    current_region = {
                        'start_row': row[0].row,
                        'headers': row_values,
                        'data': []
                    }
                elif current_region:
                    current_region['data'].append(row_values)
            elif current_region:
                regions.append(current_region)
                current_region = None
                
        if current_region:
            regions.append(current_region)
            
        return regions

    def _is_header_row(self, row_values: List[Any]) -> bool:
        """Determine if a row is likely a header row."""
        non_empty = [v for v in row_values if v is not None]
        if not non_empty:
            return False
            
        # Check if all values are strings and not too long
        return all(isinstance(v, str) and len(v) < 100 for v in non_empty)

    def _find_text_regions(self, ws) -> List[Dict[str, Any]]:
        """Find regions containing unstructured text."""
        text_regions = []
        current_text = []
        
        for row in ws.iter_rows():
            cell_values = [cell.value for cell in row if cell.value]
            if cell_values:
                # Check if this looks like unstructured text
                if any(isinstance(v, str) and len(v) > 100 for v in cell_values):
                    current_text.extend(cell_values)
                elif current_text:
                    text_regions.append({
                        'content': '\n'.join(current_text),
                        'row': row[0].row
                    })
                    current_text = []
                    
        if current_text:
            text_regions.append({
                'content': '\n'.join(current_text),
                'row': ws.max_row
            })
            
        return text_regions

class SmartExcelManager:
    def __init__(self, excel_directory: str = "./excel_files"):
        self.excel_directory = Path(excel_directory)
        self.file_cache = {}
        self.content_cache = {}
        self.analyzer = ExcelAnalyzer()
        
    def analyze_file(self, filename: str) -> Dict[str, Any]:
        """Thoroughly analyze an Excel file and cache its content structure."""
        if filename in self.content_cache:
            return self.content_cache[filename]
            
        wb = openpyxl.load_workbook(self.excel_directory / filename, data_only=True)
        analysis = {
            'worksheets': {},
            'metadata': {
                'total_sheets': len(wb.sheetnames),
                'has_charts': any(ws._charts for ws in wb.worksheets),
                'file_size': (self.excel_directory / filename).stat().st_size
            }
        }
        
        for ws in wb.worksheets:
            content = WorksheetContent(ws.title)
            content.content_type = self.analyzer.detect_content_type(ws)
            
            if content.content_type in [ContentType.STRUCTURED_TABLE, ContentType.MIXED]:
                content.data_regions = self.analyzer._find_data_regions(ws)
                
            if content.content_type in [ContentType.UNSTRUCTURED_TEXT, ContentType.MIXED]:
                content.unstructured_text = self.analyzer._find_text_regions(ws)
                
            if ws._charts:
                content.charts = [{'type': chart.type, 'anchor': chart.anchor} 
                                for chart in ws._charts]
                
            analysis['worksheets'][ws.title] = vars(content)
            
        self.content_cache[filename] = analysis
        return analysis

    def get_best_worksheet_for_query(self, filename: str, query: str) -> Tuple[str, str]:
        """Determine the best worksheet to answer a query."""
        analysis = self.analyze_file(filename)
        best_score = 0
        best_sheet = None
        best_type = None
        
        # Convert query to lowercase for matching
        query_lower = query.lower()
        
        for sheet_name, sheet_info in analysis['worksheets'].items():
            score = 0
            
            # Check sheet name relevance
            if any(word in sheet_name.lower() for word in query_lower.split()):
                score += 5
                
            # Check content type relevance
            if "graph" in query_lower or "chart" in query_lower:
                if sheet_info['charts']:
                    score += 10
                    best_type = ContentType.CHART
            elif "table" in query_lower or any(word in query_lower for word in ['sum', 'average', 'total', 'count']):
                if sheet_info['content_type'] in [ContentType.STRUCTURED_TABLE, ContentType.MIXED]:
                    score += 8
                    best_type = ContentType.STRUCTURED_TABLE
            
            # Check column headers for relevance
            for region in sheet_info.get('data_regions', []):
                headers = [str(h).lower() for h in region['headers'] if h]
                if any(word in ' '.join(headers) for word in query_lower.split()):
                    score += 7
                    
            # Check unstructured text for relevance
            for text_region in sheet_info.get('unstructured_text', []):
                if any(word in text_region['content'].lower() for word in query_lower.split()):
                    score += 6
                    best_type = ContentType.UNSTRUCTURED_TEXT
                    
            if score > best_score:
                best_score = score
                best_sheet = sheet_name
                
        return best_sheet, best_type

    def query_content(self, filename: str, query: str) -> Dict[str, Any]:
        """Smart query that adapts to content type and finds relevant information."""
        best_sheet, content_type = self.get_best_worksheet_for_query(filename, query)
        if not best_sheet:
            return {"error": "Could not find relevant worksheet"}
            
        wb = openpyxl.load_workbook(self.excel_directory / filename, data_only=True)
        ws = wb[best_sheet]
        
        if content_type == ContentType.STRUCTURED_TABLE:
            # Convert to pandas for structured analysis
            data_regions = self.analyzer._find_data_regions(ws)
            if data_regions:
                df = pd.DataFrame(data_regions[0]['data'], columns=data_regions[0]['headers'])
                try:
                    # Attempt to interpret the query as a pandas operation
                    if "average" in query.lower() or "mean" in query.lower():
                        result = df.mean(numeric_only=True).to_dict()
                    elif "sum" in query.lower() or "total" in query.lower():
                        result = df.sum(numeric_only=True).to_dict()
                    else:
                        result = df.to_dict(orient='records')
                    return {"type": "structured", "data": result}
                except Exception as e:
                    logger.error(f"Error processing structured data: {e}")
                    return {"error": str(e)}
                    
        elif content_type == ContentType.UNSTRUCTURED_TEXT:
            text_regions = self.analyzer._find_text_regions(ws)
            if text_regions:
                # Use NLP to find relevant text
                if self.analyzer.nlp:
                    relevant_texts = []
                    query_doc = self.analyzer.nlp(query)
                    for region in text_regions:
                        text_doc = self.analyzer.nlp(region['content'])
                        similarity = query_doc.similarity(text_doc)
                        if similarity > 0.5:  # Threshold for relevance
                            relevant_texts.append(region['content'])
                    return {"type": "unstructured", "data": relevant_texts}
                else:
                    # Simple keyword matching if NLP not available
                    relevant_texts = [
                        region['content'] for region in text_regions
                        if any(word.lower() in region['content'].lower() 
                              for word in query.split())
                    ]
                    return {"type": "unstructured", "data": relevant_texts}
                    
        elif content_type == ContentType.CHART:
            # Return information about charts
            charts_info = [{'type': chart.type, 'anchor': chart.anchor} 
                         for chart in ws._charts]
            return {"type": "chart", "data": charts_info}
            
        return {"error": "Could not process content type"}

class SmartExcelAgent:
    def __init__(self, model, tokenizer, excel_manager: SmartExcelManager):
        self.model = model
        self.tokenizer = tokenizer
        self.excel_manager = excel_manager
        
    async def process_query(self, query: str) -> Dict[str, Any]:
        """Process a query about Excel files."""
        try:
            # First, list available files
            available_files = [f.name for f in self.excel_directory.glob("*.xlsx")]
            
            # Analyze query to determine relevant files
            relevant_files = []
            for file in available_files:
                analysis = self.excel_manager.analyze_file(file)
                # Use the model to determine relevance
                relevance_prompt = f"Is the file '{file}' relevant to the query: '{query}'?"
                # (Add relevance checking logic here)
                
            if not relevant_files:
                # If no specific files mentioned, search all files
                relevant_files = available_files
                
            results = []
            for file in relevant_files:
                result = self.excel_manager.query_content(file, query)
                if 'error' not in result:
                    results.append({
                        'file': file,
                        'result': result
                    })
                    
            return {
                'status': 'success',
                'results': results,
                'files_checked': len(relevant_files)
            }
            
        except Exception as e:
            logger.error(f"Error processing query: {e}")
            return {
                'status': 'error',
                'message': str(e)
            }

# Initialize the system
excel_manager = SmartExcelManager()
agent = SmartExcelAgent(model, tokenizer, excel_manager)

@app.post("/query")
async def query_model(request: QueryRequest):
    return await agent.process_query(request.query)